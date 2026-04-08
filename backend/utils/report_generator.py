"""Report generation utilities"""
import json
import logging
from datetime import datetime
from typing import Dict, Any
from pathlib import Path
import os
from xml.sax.saxutils import escape as xml_escape

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
    from reportlab.lib import colors
except ImportError:
    # Handle missing reportlab gracefully
    pass

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate competitive intelligence reports in multiple formats"""
    
    def __init__(self, output_dir: str = None):
        if output_dir is None:
            base_dir = Path(__file__).resolve().parent.parent
            output_dir = str(base_dir / 'reports')
        self.output_dir = output_dir
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)

    def _clean_list(self, values):
        if not isinstance(values, list):
            return []
        return [str(value).strip() for value in values if str(value).strip()]

    def _safe_text(self, value):
        return xml_escape(str(value or 'N/A'))

    def _format_joined(self, values, fallback='N/A'):
        cleaned = self._clean_list(values)
        return ', '.join(cleaned) if cleaned else fallback

    def _brand_slug(self, brand_name: str) -> str:
        return str(brand_name or '').strip().lower().replace(' ', '_')

    def _cleanup_previous_brand_reports(self, brand_name: str) -> None:
        """Remove older generated files for the same brand so only latest remains."""
        slug = self._brand_slug(brand_name)
        if not slug:
            return
        prefix = f"{slug}_report_"
        try:
            for filename in os.listdir(self.output_dir):
                if filename.startswith(prefix):
                    file_path = os.path.join(self.output_dir, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
        except Exception as cleanup_exc:
            logger.warning(f"Could not fully clean previous reports for {brand_name}: {cleanup_exc}")

    def _cleanup_all_reports(self) -> None:
        """Remove all previously generated report files before a fresh run."""
        try:
            for filename in os.listdir(self.output_dir):
                if '_report_' in filename:
                    file_path = os.path.join(self.output_dir, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
        except Exception as cleanup_exc:
            logger.warning(f"Could not fully clean reports directory: {cleanup_exc}")

    def _build_competitor_section(self, brand_name: str, competitor: Dict[str, Any]) -> Dict[str, Any]:
        insights = competitor.get('insights', {}) or {}
        sentiment = insights.get('customer_sentiment', {}) or {}
        pricing = insights.get('pricing_strategy', {}) or {}
        key_features = self._clean_list(insights.get('key_features', []))
        messaging_themes = self._clean_list(insights.get('messaging_themes', []))
        market_gaps = self._clean_list(insights.get('market_gaps', []))
        opportunity_areas = self._clean_list(insights.get('opportunity_areas', []))
        competitive_threats = self._clean_list(insights.get('competitive_threats', []))
        positives = self._clean_list(sentiment.get('positives', []))
        negatives = self._clean_list(sentiment.get('negatives', []))

        positioning = messaging_themes[0] if messaging_themes else f"{competitor.get('competitor_name', 'The competitor')} is shaping the category narrative."

        recent_feature_launches = []
        for feature in key_features[:3]:
            recent_feature_launches.append({
                'title': feature,
                'strategic_implication': (
                    f"Signals {competitor.get('competitor_name', 'the competitor')}'s focus on {feature.lower()} and the pressure it creates for {brand_name}."
                )
            })

        customer_love = positives or [
            f"Buyers value {competitor.get('competitor_name', 'the competitor')}'s core offering and category recognition."
        ]
        customer_hate = negatives or [
            f"Users report friction around complexity and pricing tradeoffs."
        ]

        pricing_description = pricing.get('model') or 'Pricing details were not fully available from the source data.'
        if pricing.get('tiers') or pricing.get('price_range'):
            pricing_description = (
                f"{pricing_description} | Tiers: {self._format_joined(pricing.get('tiers'), 'N/A')} | "
                f"Range: {pricing.get('price_range', 'N/A')}"
            )

        return {
            'competitor_name': competitor.get('competitor_name', 'Unknown'),
            'positioning': positioning,
            'recent_feature_launches': recent_feature_launches,
            'strengths': customer_love,
            'what_customers_love': customer_love,
            'what_customers_hate': customer_hate,
            'pricing': pricing_description,
            'weaknesses': market_gaps or [
                f"No explicit weaknesses were returned for {competitor.get('competitor_name', 'the competitor')}."
            ],
            'notion_opportunities': [
                (
                    f"{brand_name} can win by addressing '{weakness}' with simpler onboarding, clearer value communication, and stronger execution."
                )
                for weakness in (market_gaps[:3] or ['General category friction'])
            ],
            'weaknesses_and_opportunities': [
                {
                    'weakness': weakness,
                    'opportunity': (
                        f"{brand_name} can contrast this by emphasizing clarity, ease of adoption, and measurable ROI against {competitor.get('competitor_name', 'the competitor')}."
                    )
                }
                for weakness in (market_gaps[:3] or ["Competitive overlap"])
            ],
            'opportunity_areas': opportunity_areas,
            'competitive_threats': competitive_threats,
        }

    def _build_market_strategy(self, brand_name: str, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        market_analysis = analysis_data.get('market_analysis', {}) or {}
        competitor_data = analysis_data.get('competitor_data', []) or []
        campaign = analysis_data.get('campaign_recommendations', {}) or {}

        biggest_threat = market_analysis.get('key_threats', [
            f"Category pressure from established competitors in the market"
        ])[0]
        biggest_opportunity = market_analysis.get('opportunities', [
            f"Win buyers with a clearer and easier-to-adopt alternative"
        ])[0]

        whitespace = (
            f"No competitor is fully owning the narrative of being both powerful and easy to adopt. "
            f"{brand_name} can occupy that gap with a stronger onboarding story and simpler value communication."
        )

        competitor_vulnerabilities = []
        for competitor in competitor_data[:3]:
            weaknesses = self._clean_list(competitor.get('insights', {}).get('market_gaps', []))
            if weaknesses:
                competitor_vulnerabilities.append({
                    'competitor': competitor.get('competitor_name', 'Unknown'),
                    'vulnerability': weaknesses[0],
                    'action': f"Use this weakness as a message point in {brand_name}'s positioning and campaign copy."
                })

        if not competitor_vulnerabilities:
            competitor_vulnerabilities.append({
                'competitor': 'Category leaders',
                'vulnerability': 'Complexity and switching friction',
                'action': f"Make simplicity and faster time-to-value the core contrast for {brand_name}."
            })

        return {
            'biggest_threat': biggest_threat,
            'biggest_opportunity': biggest_opportunity,
            'market_whitespace': whitespace,
            'market_trend': market_analysis.get('market_trends', [
                'AI-driven productivity and all-in-one workspace consolidation'
            ])[0],
            'competitor_vulnerabilities_to_exploit': competitor_vulnerabilities,
            'recommended_positioning': {
                'recommended_angle': campaign.get('overall_strategy', f'{brand_name}: intuitive, high-value, and easier to adopt.'),
                'rationale': campaign.get('overall_strategy', 'Position around clarity, ease of use, and measurable outcomes.'),
                'proof_points': self._clean_list(campaign.get('messaging_positioning', {}).get('value_props', []))
            }
        }

    def _build_run_comparison(self, brand_name: str, analysis_data: Dict[str, Any], previous_analysis: Dict[str, Any] = None) -> Dict[str, Any]:
        def signal_texts(competitor_items):
            texts = []
            for competitor in competitor_items:
                for signal in competitor.get('signals', []):
                    if isinstance(signal, dict):
                        signal_text = signal.get('text') or signal.get('signal_text')
                        if signal_text:
                            texts.append(str(signal_text).strip())
                    elif signal:
                        texts.append(str(signal).strip())
            return [text for text in texts if text]

        current_competitors = [
            competitor.get('competitor_name')
            for competitor in analysis_data.get('competitor_data', [])
            if competitor.get('competitor_name')
        ]

        current_signals = signal_texts(analysis_data.get('competitor_data', []))

        if not previous_analysis:
            return {
                'comparison': 'Initial run for this brand. No earlier report was available for comparison.',
                'new_signals': current_signals[:6],
                'removed_signals': [],
                'competitor_changes': {
                    'added': current_competitors,
                    'removed': []
                }
            }

        previous_competitors = [
            competitor.get('competitor_name')
            for competitor in previous_analysis.get('competitor_data', [])
            if competitor.get('competitor_name')
        ]
        previous_signals = signal_texts(previous_analysis.get('competitor_data', []))

        added_competitors = [name for name in current_competitors if name not in previous_competitors]
        removed_competitors = [name for name in previous_competitors if name not in current_competitors]
        added_signals = [signal for signal in current_signals if signal not in previous_signals]
        removed_signals = [signal for signal in previous_signals if signal not in current_signals]

        current_threat = (analysis_data.get('market_analysis', {}) or {}).get('threat_level', 'N/A')
        previous_threat = (previous_analysis.get('market_analysis', {}) or {}).get('threat_level', 'N/A')

        summary_parts = []
        if added_competitors:
            summary_parts.append(f"New competitors surfaced: {', '.join(added_competitors)}.")
        if removed_competitors:
            summary_parts.append(f"Competitors no longer highlighted: {', '.join(removed_competitors)}.")
        if current_threat != previous_threat:
            summary_parts.append(f"Threat level changed from {previous_threat} to {current_threat}.")
        if not summary_parts:
            summary_parts.append('The core competitor set stayed stable, but the messaging and opportunity framing were refreshed.')

        return {
            'comparison': f"Comparing: {previous_analysis.get('analyzed_at', 'previous run')} → {analysis_data.get('analyzed_at', 'current run')}",
            'summary': ' '.join(summary_parts),
            'new_signals': added_signals[:8],
            'removed_signals': removed_signals[:8],
            'competitor_changes': {
                'added': added_competitors,
                'removed': removed_competitors
            }
        }
    
    def generate_json_report(self, brand_name: str, analysis_data: Dict[str, Any], previous_analysis: Dict[str, Any] = None) -> str:
        """Generate a structured JSON report."""
        filename = f"{brand_name.lower().replace(' ', '_')}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(self.output_dir, filename)

        competitor_sections = [
            self._build_competitor_section(brand_name, competitor)
            for competitor in analysis_data.get('competitor_data', [])
        ]
        market_strategy = self._build_market_strategy(brand_name, analysis_data)
        run_comparison = self._build_run_comparison(brand_name, analysis_data, previous_analysis)
        campaign = analysis_data.get('campaign_recommendations', {}) or {}
        guardrails = analysis_data.get('guardrails', {}) or {}
        guardrail_summary = {
            'version': guardrails.get('version', 'unknown'),
            'brand_name': guardrails.get('brand_name', brand_name),
            'category': guardrails.get('category', 'unknown'),
            'input_candidates_count': guardrails.get('input_candidates_count', 0),
            'approved_count': guardrails.get('approved_count', 0),
            'rejected_count': guardrails.get('rejected_count', 0),
            'approved_competitors': guardrails.get('approved_competitors', []),
            'rejected_competitors': guardrails.get('rejected_competitors', []),
            'final_selected_competitors': guardrails.get('final_selected_competitors', []),
            'llm_validation_used': guardrails.get('checks', {}).get('llm_relevance_validation_used', False),
        }
        
        report = {
            'report_title': f'{brand_name} Competitive Intelligence Report',
            'generated_at': datetime.utcnow().isoformat(),
            'brand_analyzed': brand_name,
            'status': 'Fresh data — ready to act on',
            'analysis_summary': {
                'competitors_analyzed': analysis_data.get('competitors_analyzed', 0),
                'signals_extracted': analysis_data.get('signals_extracted', 0),
                'market_threats': len(analysis_data.get('market_analysis', {}).get('key_threats', [])),
                'opportunities_identified': len(analysis_data.get('market_analysis', {}).get('opportunities', []))
            },
            'competitive_intelligence_breakdown': competitor_sections,
            'market_strategic_analysis': market_strategy,
            'campaign_recommendations': campaign,
            'guardrail_summary': guardrail_summary,
            'what_changed_since_last_run': run_comparison,
            'footer': {
                'generated_by': 'KeenFox Competitive Intelligence System',
                'model': 'Gemini',
                'refresh_hint': f'POST /api/analyze with brand_name="{brand_name}"',
                'query_hint': f'GET /api/reports or use the dashboard for {brand_name}'
            },
            'detailed_analysis': analysis_data
        }
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2)
            logger.info(f"Generated JSON report: {filepath}")
            return filepath
        except Exception as write_exc:
            logger.error(f"Error writing JSON report for {brand_name}: {write_exc}")
            return None
    
    def generate_markdown_report(self, brand_name: str, analysis_data: Dict[str, Any]) -> str:
        """Generate Markdown report for easy reading"""
        filename = f"{brand_name.lower().replace(' ', '_')}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.md"
        filepath = os.path.join(self.output_dir, filename)
        
        md_content = f"""# Competitive Intelligence Report - {brand_name}

**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}

## Executive Summary

- **Competitors Analyzed:** {analysis_data.get('competitors_analyzed', 0)}
- **Signals Extracted:** {analysis_data.get('signals_extracted', 0)}
- **Report Date:** {datetime.utcnow().strftime('%Y-%m-%d')}

---

## Market Analysis

### Market Position
**Position:** {analysis_data.get('market_analysis', {}).get('market_position', 'N/A')}

**Threat Level:** {analysis_data.get('market_analysis', {}).get('threat_level', 'N/A')}

### Key Threats
"""
        
        for threat in analysis_data.get('market_analysis', {}).get('key_threats', []):
            md_content += f"- {threat}\n"
        
        md_content += "\n### Opportunities\n"
        for opp in analysis_data.get('market_analysis', {}).get('opportunities', []):
            md_content += f"- {opp}\n"
        
        md_content += "\n---\n## Competitive Landscape\n\n"
        
        for competitor in analysis_data.get('competitor_data', []):
            md_content += f"### {competitor.get('competitor_name')}\n\n"
            md_content += f"**Threat Level:** {competitor.get('threat_level', 'N/A')}\n\n"
            md_content += "**Key Features:**\n"
            for feature in competitor.get('insights', {}).get('key_features', [])[:5]:
                md_content += f"- {feature}\n"
            
            md_content += "\n**Weaknesses:**\n"
            for weakness in competitor.get('insights', {}).get('market_gaps', [])[:3]:
                md_content += f"- {weakness}\n"
            
            md_content += "\n"

        guardrails = analysis_data.get('guardrails', {}) or {}
        if guardrails:
            md_content += "\n---\n## Guardrails Summary\n\n"
            md_content += f"- **Approved Competitors:** {guardrails.get('approved_count', 0)}\n"
            md_content += f"- **Rejected Candidates:** {guardrails.get('rejected_count', 0)}\n"
            md_content += f"- **Final Selected Competitors:** {', '.join(guardrails.get('final_selected_competitors', [])) or 'N/A'}\n"
            md_content += f"- **LLM Validation Used:** {'Yes' if guardrails.get('checks', {}).get('llm_relevance_validation_used') else 'No'}\n"
            md_content += "\n### Rejected Candidates\n"
            for item in guardrails.get('rejected_competitors', []):
                md_content += f"- {item.get('name', 'Unknown')}: {', '.join(item.get('reasons', [])) or 'guardrail rejection'}\n"
        
        md_content += "\n---\n## Campaign Recommendations\n\n"
        
        camp_rec = analysis_data.get('campaign_recommendations', {})
        
        md_content += f"### Overall Strategy\n{camp_rec.get('overall_strategy', 'N/A')}\n\n"
        
        md_content += "### Messaging & Positioning\n"
        messaging = camp_rec.get('messaging_positioning', {})
        md_content += f"**Headline:** {messaging.get('headline', 'N/A')}\n\n"
        md_content += f"**Differentiation:** {messaging.get('differentiation', 'N/A')}\n\n"
        
        md_content += "### Channel Strategy\n"
        channel = camp_rec.get('channel_strategy', {})
        md_content += f"**Primary Channels:**\n"
        for ch in channel.get('primary_channels', []):
            md_content += f"- {ch}\n"
        
        md_content += f"\n**Rationale:** {channel.get('rationale', 'N/A')}\n\n"
        
        md_content += "### GTM Recommendations\n"
        for i, rec in enumerate(camp_rec.get('gtm_recommendations', [])[:5], 1):
            md_content += f"\n#### {i}. {rec.get('title', 'N/A')} ({rec.get('priority', 'N/A')} Priority)\n"
            md_content += f"**Description:** {rec.get('description', 'N/A')}\n\n"
            md_content += f"**Rationale:** {rec.get('rationale', 'N/A')}\n\n"
            md_content += f"**Expected Impact:** {rec.get('expected_impact', 'N/A')}\n\n"
            md_content += f"**Timeline:** {rec.get('timeline', 'N/A')}\n"
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(md_content)
            logger.info(f"Generated Markdown report: {filepath}")
            return filepath
        except Exception as write_exc:
            logger.error(f"Error writing Markdown report for {brand_name}: {write_exc}")
            return None
    
    def generate_pdf_report(self, brand_name: str, analysis_data: Dict[str, Any], previous_analysis: Dict[str, Any] = None) -> str:
        """Generate a document-style PDF report."""
        try:
            filename = f"{brand_name.lower().replace(' ', '_')}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(self.output_dir, filename)
            
            doc = SimpleDocTemplate(
                filepath,
                pagesize=letter,
                rightMargin=36,
                leftMargin=36,
                topMargin=42,
                bottomMargin=36,
            )
            story = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontName='Times-Bold',
                fontSize=14,
                leading=16,
                textColor=colors.black,
                spaceAfter=6,
            )
            section_style = ParagraphStyle(
                'SectionHeading',
                parent=styles['Heading2'],
                fontName='Times-Bold',
                fontSize=14,
                leading=16,
                textColor=colors.black,
                spaceBefore=12,
                spaceAfter=8,
            )
            sub_style = ParagraphStyle(
                'SubHeading',
                parent=styles['Heading3'],
                fontName='Times-Bold',
                fontSize=12,
                leading=14,
                textColor=colors.black,
                spaceBefore=6,
                spaceAfter=4,
            )
            body_style = ParagraphStyle(
                'Body',
                parent=styles['BodyText'],
                fontName='Times-Roman',
                fontSize=12,
                leading=14,
                textColor=colors.black,
            )
            small_style = ParagraphStyle(
                'Small',
                parent=styles['BodyText'],
                fontName='Times-Roman',
                fontSize=12,
                leading=14,
                textColor=colors.black,
            )

            note_style = ParagraphStyle(
                'Note',
                parent=styles['BodyText'],
                fontSize=9,
                leading=12,
                textColor=colors.HexColor('#2a3d52'),
            )

            accent_style = ParagraphStyle(
                'Accent',
                parent=styles['BodyText'],
                fontSize=9,
                leading=12,
                textColor=colors.HexColor('#0b6d5b'),
            )

            market_analysis = analysis_data.get('market_analysis', {})
            campaign = analysis_data.get('campaign_recommendations', {})
            competitors = analysis_data.get('competitor_data', [])
            market_strategy = self._build_market_strategy(brand_name, analysis_data)
            run_comparison = self._build_run_comparison(brand_name, analysis_data, previous_analysis)

            def bullet_list(items, fallback_text='No direct customer signal was captured; validate this point with fresh reviews and competitor release notes.'):
                cleaned = self._clean_list(items)
                if not cleaned:
                    return Paragraph(self._safe_text(fallback_text), body_style)
                text = '<br/>'.join([f'• {self._safe_text(item)}' for item in cleaned])
                return Paragraph(text, body_style)

            def wrap_cell(value, style=body_style):
                return Paragraph(self._safe_text(value), style)

            def table_style(table):
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
                    ('FONTSIZE', (0, 1), (-1, -1), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('WORDWRAP', (0, 0), (-1, -1), 'LTR'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                return table

            def cover_style(table):
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('BOX', (0, 0), (-1, -1), 1, colors.black),
                    ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
                    ('FONTNAME', (0, 0), (0, -1), 'Times-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 12),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                    ('TOPPADDING', (0, 0), (-1, -1), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
                ]))
                return table

            # Cover section
            story.append(Paragraph('KeenFox Competitive Intelligence Report', title_style))
            story.append(Paragraph(self._safe_text(brand_name), section_style))
            story.append(Spacer(1, 8))
            story.append(Paragraph(
                f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
                small_style
            ))
            story.append(Spacer(1, 10))
            story.append(cover_style(Table([
                [wrap_cell('Brand Analyzed', sub_style), wrap_cell(brand_name)],
                [wrap_cell('Status', sub_style), wrap_cell('Fresh data — ready to act on')],
                [wrap_cell('Competitors Analyzed', sub_style), wrap_cell(str(analysis_data.get('competitors_analyzed', 0)))],
                [wrap_cell('Signals Extracted', sub_style), wrap_cell(str(analysis_data.get('signals_extracted', 0)))],
            ], colWidths=[160, 340])))
            story.append(Spacer(1, 10))
            story.append(Spacer(1, 4))
            story.append(Paragraph(
                '1. Executive Summary',
                section_style
            ))
            story.append(Paragraph(
                campaign.get('overall_strategy', f'Competitive intelligence summary for {brand_name}.'),
                body_style
            ))
            story.append(Spacer(1, 10))

            summary_table = table_style(Table([
                [wrap_cell('Metric', sub_style), wrap_cell('Value', sub_style)],
                [wrap_cell('Competitors Analyzed', sub_style), wrap_cell(str(analysis_data.get('competitors_analyzed', 0)))],
                [wrap_cell('Signals Extracted', sub_style), wrap_cell(str(analysis_data.get('signals_extracted', 0)))],
                [wrap_cell('Market Threats', sub_style), wrap_cell(str(len(market_analysis.get('key_threats', []))))],
                [wrap_cell('Opportunities Identified', sub_style), wrap_cell(str(len(market_analysis.get('opportunities', []))))],
            ], colWidths=[170, 330]))
            story.append(summary_table)
            story.append(Spacer(1, 14))

            # Competitive breakdown section
            story.append(Paragraph('2. Competitive Intelligence Breakdown', section_style))
            for idx, competitor in enumerate(competitors, 1):
                insights = competitor.get('insights', {})
                pricing_strategy = insights.get('pricing_strategy', {}) or {}
                pricing_text = pricing_strategy.get('model') or 'Pricing details were not fully available from the source data.'
                if pricing_strategy.get('tiers') or pricing_strategy.get('price_range'):
                    pricing_text = (
                        f"{pricing_text} | Tiers: {self._format_joined(pricing_strategy.get('tiers'), 'N/A')} | "
                        f"Range: {pricing_strategy.get('price_range', 'N/A')}"
                    )
                story.append(Paragraph(f'2.{idx} {self._safe_text(competitor.get("competitor_name", "Unknown"))}', sub_style))
                story.append(Paragraph(
                    f"<b>Positioning:</b> {self._safe_text(self._format_joined(insights.get('messaging_themes', [])[:2], 'N/A'))}",
                    body_style
                ))
                story.append(Paragraph('Recent Feature Launches', sub_style))
                launch_lines = []
                for feature in self._clean_list(insights.get('key_features', []))[:3]:
                    launch_lines.append(f'• {self._safe_text(feature)} - Strategic implication: {self._safe_text(competitor.get("competitor_name", "The competitor"))} is signaling product depth and forcing the market to respond.')
                story.append(Paragraph('<br/>'.join(launch_lines) if launch_lines else self._safe_text('No launch signal was captured in this run; monitor release notes and changelogs for updates.'), body_style))
                story.append(Paragraph('Strengths (What Customers Love)', sub_style))
                story.append(bullet_list(insights.get('customer_sentiment', {}).get('positives', [])))
                story.append(Paragraph('Weaknesses (What Customers Hate) / Notion Opportunities', sub_style))
                hate_lines = []
                for item in self._clean_list(insights.get('customer_sentiment', {}).get('negatives', []))[:4]:
                    hate_lines.append(f'• {self._safe_text(item)}')
                if not hate_lines:
                    hate_lines.append('• Limited negative sentiment was available in the source data.')
                story.append(Paragraph('<br/>'.join(hate_lines), body_style))
                story.append(Paragraph(f"<b>Pricing:</b> {self._safe_text(pricing_text)}", body_style))
                story.append(Paragraph('Weaknesses & Opportunities', sub_style))
                weakness_lines = []
                for weakness in self._clean_list(insights.get('market_gaps', []))[:3]:
                    weakness_lines.append(
                        f'• Weakness: {self._safe_text(weakness)} → {self._safe_text(brand_name)} can emphasize simpler onboarding, clearer ROI, and cleaner execution.'
                    )
                story.append(Paragraph('<br/>'.join(weakness_lines) if weakness_lines else self._safe_text(f'• Weakness: Limited direct weakness data in this run → {brand_name} can validate via customer interviews and comparison-page search queries.'), body_style))
                story.append(Spacer(1, 8))

            story.append(PageBreak())

            story.append(Paragraph('3. Market Strategic Analysis', section_style))
            story.append(Paragraph(f"<b>Biggest Threat:</b> {self._safe_text(market_strategy['biggest_threat'])}", body_style))
            story.append(Spacer(1, 4))
            story.append(Paragraph(f"<b>Biggest Opportunity:</b> {self._safe_text(market_strategy['biggest_opportunity'])}", body_style))
            story.append(Spacer(1, 4))
            story.append(Paragraph(f"<b>Market Whitespace:</b> {self._safe_text(market_strategy['market_whitespace'])}", body_style))
            story.append(Spacer(1, 4))
            story.append(Paragraph(f"<b>Market Trend:</b> {self._safe_text(market_strategy['market_trend'])}", body_style))
            story.append(Spacer(1, 8))
            story.append(Paragraph('Competitor Vulnerabilities to Exploit', sub_style))
            vulnerability_rows = []
            for item in market_strategy['competitor_vulnerabilities_to_exploit']:
                vulnerability_rows.append([
                    self._safe_text(item.get('competitor', 'Unknown')),
                    self._safe_text(item.get('vulnerability', 'N/A')),
                    self._safe_text(item.get('action', 'N/A'))
                ])
            vulnerability_table_rows = [[
                wrap_cell('Competitor', sub_style),
                wrap_cell('Vulnerability', sub_style),
                wrap_cell('Action', sub_style)
            ]]
            for row in vulnerability_rows:
                vulnerability_table_rows.append([
                    wrap_cell(row[0]),
                    wrap_cell(row[1]),
                    wrap_cell(row[2])
                ])
            story.append(table_style(Table(vulnerability_table_rows, colWidths=[100, 140, 230])))
            story.append(Spacer(1, 10))
            story.append(Paragraph('Recommended Positioning', sub_style))
            story.append(Paragraph(
                f"<b>Recommended Angle:</b> {self._safe_text(market_strategy['recommended_positioning']['recommended_angle'])}",
                body_style
            ))
            story.append(Paragraph(
                f"<b>Rationale:</b> {self._safe_text(market_strategy['recommended_positioning']['rationale'])}",
                body_style
            ))
            proof_points = market_strategy['recommended_positioning'].get('proof_points', [])
            story.append(Paragraph('Proof Points KeenFox Needs', sub_style))
            story.append(bullet_list(proof_points or [
                'Demonstrable improvements in onboarding flows and time-to-first-value metrics.',
                'Clearer product templates and faster activation stories.',
                'Performance benchmarks for large workspaces and data-heavy use cases.'
            ]))

            story.append(PageBreak())

            story.append(Paragraph('4. Campaign Recommendations', section_style))
            story.append(Paragraph('What Changed Since Last Run', sub_style))
            story.append(Paragraph(self._safe_text(run_comparison['comparison']), body_style))
            story.append(Paragraph(self._safe_text(run_comparison.get('summary', '')), body_style))
            story.append(Paragraph('New Signals', sub_style))
            story.append(bullet_list(run_comparison.get('new_signals', [])))
            story.append(Paragraph('Removed Signals', sub_style))
            story.append(bullet_list(run_comparison.get('removed_signals', [])))
            story.append(Spacer(1, 8))

            guardrails = analysis_data.get('guardrails', {}) or {}
            if guardrails:
                story.append(Paragraph('Guardrails Summary', sub_style))
                guardrail_rows = [
                    ['Approved Competitors', str(guardrails.get('approved_count', 0))],
                    ['Rejected Candidates', str(guardrails.get('rejected_count', 0))],
                    ['Final Selected Competitors', ', '.join(guardrails.get('final_selected_competitors', [])) or 'N/A'],
                    ['LLM Validation Used', 'Yes' if guardrails.get('checks', {}).get('llm_relevance_validation_used') else 'No'],
                ]
                guardrail_table_rows = [[wrap_cell('Field', sub_style), wrap_cell('Value', sub_style)]]
                for row in guardrail_rows:
                    guardrail_table_rows.append([wrap_cell(row[0], sub_style), wrap_cell(row[1])])
                story.append(table_style(Table(guardrail_table_rows, colWidths=[150, 350])))
                story.append(Spacer(1, 8))

                rejected = guardrails.get('rejected_competitors', [])
                if rejected:
                    story.append(Paragraph('Rejected Candidates', sub_style))
                    for item in rejected:
                        story.append(Paragraph(
                            f"<b>{self._safe_text(item.get('name', 'Unknown'))}</b> - {self._safe_text(', '.join(item.get('reasons', [])) or 'guardrail rejection')}",
                            body_style
                        ))
                    story.append(Spacer(1, 6))

            messaging = campaign.get('messaging_positioning', {})
            channel = campaign.get('channel_strategy', {})
            gtm_recommendations = campaign.get('gtm_recommendations', [])

            story.append(Paragraph('Messaging & Positioning', sub_style))
            messaging_rows = [
                ['Headline', messaging.get('headline', 'N/A')],
                ['Subheadline', messaging.get('subheadline', 'N/A')],
                ['Differentiation', messaging.get('differentiation', 'N/A')],
            ]
            messaging_table_rows = [[wrap_cell('Field', sub_style), wrap_cell('Copy', sub_style)]]
            for row in messaging_rows:
                messaging_table_rows.append([wrap_cell(row[0], sub_style), wrap_cell(row[1])])
            story.append(table_style(Table(messaging_table_rows, colWidths=[120, 380])))
            story.append(Spacer(1, 8))

            story.append(Paragraph('Channel Strategy', sub_style))
            channel_rows = [
                ['Primary Channels', ', '.join(channel.get('primary_channels', [])) or 'N/A'],
                ['Secondary Channels', ', '.join(channel.get('secondary_channels', [])) or 'N/A'],
                ['Rationale', channel.get('rationale', 'N/A')],
            ]
            channel_table_rows = [[wrap_cell('Field', sub_style), wrap_cell('Details', sub_style)]]
            for row in channel_rows:
                channel_table_rows.append([wrap_cell(row[0], sub_style), wrap_cell(row[1])])
            story.append(table_style(Table(channel_table_rows, colWidths=[130, 370])))
            story.append(Spacer(1, 10))

            story.append(Paragraph('GTM Recommendations', sub_style))
            for idx, rec in enumerate(gtm_recommendations[:5], 1):
                story.append(Paragraph(f'{idx}. {rec.get("title", "N/A")} ({rec.get("priority", "N/A")})', body_style))
                rec_rows = [
                    ['Description', rec.get('description', 'N/A')],
                    ['Rationale', rec.get('rationale', 'N/A')],
                    ['Expected Impact', rec.get('expected_impact', 'N/A')],
                    ['Timeline', rec.get('timeline', 'N/A')],
                ]
                rec_table_rows = [[wrap_cell('Field', sub_style), wrap_cell('Detail', sub_style)]]
                for row in rec_rows:
                    rec_table_rows.append([wrap_cell(row[0], sub_style), wrap_cell(row[1])])
                story.append(table_style(Table(rec_table_rows, colWidths=[120, 380])))
                story.append(Spacer(1, 8))

            story.append(Paragraph('5. Footer', section_style))
            story.append(Paragraph('Report Footer', sub_style))
            story.append(Paragraph(
                f"Report generated by KeenFox Competitive Intelligence System | Brand: {self._safe_text(brand_name)} | Model: Gemini",
                small_style
            ))
            story.append(Paragraph(
                f"To refresh: POST /api/analyze with brand_name=\"{self._safe_text(brand_name)}\" | To query: GET /api/reports",
                small_style
            ))

            doc.build(story)
            logger.info(f"Generated PDF report: {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")
            return None
    
    def generate_excel_report(self, brand_name: str, analysis_data: Dict[str, Any]) -> str:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            filename = f"{brand_name.lower().replace(' ', '_')}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Header
            ws['A1'] = f"Competitive Intelligence Report - {brand_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A4'] = "Summary"
            ws['A4'].font = Font(bold=True)
            ws['A5'] = "Competitors Analyzed"
            ws['B5'] = analysis_data.get('competitors_analyzed', 0)
            ws['A6'] = "Signals Extracted"
            ws['B6'] = analysis_data.get('signals_extracted', 0)
            
            # Competitors
            ws['A8'] = "Competitors"
            ws['A8'].font = Font(bold=True)
            
            row = 9
            for competitor in analysis_data.get('competitor_data', [])[:10]:
                ws[f'A{row}'] = competitor.get('competitor_name')
                ws[f'B{row}'] = competitor.get('threat_level', 'N/A')
                row += 1
            
            wb.save(filepath)
            logger.info(f"Generated Excel report: {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return None
    
    def generate_all_formats(self, brand_name: str, analysis_data: Dict[str, Any], previous_analysis: Dict[str, Any] = None) -> Dict[str, str]:
        """Generate reports in all supported formats"""
        reports = {}
        
        # JSON
        reports['json'] = self.generate_json_report(brand_name, analysis_data, previous_analysis=previous_analysis)
        
        # PDF (if available)
        pdf_path = self.generate_pdf_report(brand_name, analysis_data, previous_analysis=previous_analysis)
        if pdf_path:
            reports['pdf'] = pdf_path
        
        return reports
